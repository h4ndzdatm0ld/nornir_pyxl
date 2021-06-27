"""Helper Functions."""
from pathlib import Path
import datetime
from openpyxl import load_workbook


def standardize(some_str):
    """Performs several string manipulations to attempt to standardize dict keys.

    Args:
        some_str (str): Manipulated string
    """
    return some_str.strip().replace(" ", "_").replace("-", "_").replace("'", "").lower()


def check_value(value):
    """Checks the contents of the value passed into the function. If condition isn't met just return the same value.
    Expect this to expand as bugs arise.

    Args:
        value (Any): Any type of value found within a row

    Returns:
        str: datetime object casted to str to form proper JSON
    """
    if isinstance(value, datetime.datetime):
        # Maybe add more logic here and standardize.
        # 'date': '2014-01-01 00:00:00', - > Looks terrible
        return str(value)
    if isinstance(value, str):
        # If value is actually an interger, return as such.
        if value.isdigit():
            return int(value)
    if value in ("None", "none"):
        return False
    return value


def all_false(some_dict):
    """Check all values of a dictionary are not False.

    Args:
        some_dict (dict): A dictionary that will be evaluated against the condition.

    Returns:
        bool: True or False depending on ALL values.
    """
    return all(value is False for value in some_dict.values())


def _calculate_reset(sheet_obj):
    """Calculate and Reset the Sheet Dimensions.

    Args:
        sheet_obj (openpyxl.worksheet): Sheet Object

    Return:
        sheet (openpyxl.worksheet): Recalculated Sheet Object
    """
    # Force Calculate Dimensions

    sheet_obj.calculate_dimension(force=True)
    # Reset Dimensions
    sheet_obj.reset_dimensions()
    # Recalculate Dimensions
    sheet_obj.calculate_dimension(force=True)

    return sheet_obj


def _check_file(file_name):
    """Check file_name exists based on input."""
    file_path = Path(file_name)

    if not file_name.endswith(".xlsx"):
        raise ValueError(f"{file_path} must end with 'xlsx'.")
    return file_path.exists()


def open_excel_wb(file_name, sheetname):
    """Load an Excel Workbook.

    Args:
        file_name (str): File Path to Excel Workbook
        sheetname (str): Sheet Name to load

    Returns:
        Sheetname (openpyxl.worksheet): Loaded OpenPyxl Sheet Object
    """
    # Check if file provided via '-f' exists.
    # If it doesn't, exit the program.
    if not _check_file(file_name):
        raise ValueError(f"{file_name} does not exist.")
    # Load the workbook
    workbook = load_workbook(
        filename=file_name, read_only=True, keep_vba=False, data_only=True
    )
    # Ensure sheetname exists inside the WorkBook.
    if sheetname not in workbook.sheetnames:
        raise ValueError(f"{sheetname} does not exist.")
    # openpyxl.worksheet._read_only.ReadOnlyWorksheet
    # Recalculate Dimensions
    return _calculate_reset(workbook[sheetname])
