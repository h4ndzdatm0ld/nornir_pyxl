from nornir.core.task import Result, Task
from openpyxl import load_workbook


def wb_sdata(
    task: Task,
    workbook: str,
    sheetname: str,
    data_only: bool = True,
    keep_vba: bool = True,
) -> Result:

    """Loads a specific sheet from a workbook(xlsx file).

    Creates a list of dictionaries using the first row as the keys.

    Arguements:

        workbookfile: Full path to .xlsx file\n
        sheetname: Worksheet Name\n
        data_only: Boolean\n
        keep_vba: Boolean\n

    Examples:

        nr.run(task=wb_sdata, workbookfile="example-wb.xlsx',
               sheetname='ip_data')
    Returns:

        Result object with the following attributes set:
        * result (''list''): list of dictionaries with
        data from the specific worksheet within the workbook.

    Notes:

        There are several flags that can be used in load_workbook.

        data_only: controls whether cells with formulas have either the formula (default) or the value stored the last time Excel read the sheet.
        keep_vba: controls whether any Visual Basic elements are preserved or not (default). If they are preserved they are still not editable.
        read_only: This is hardcoded set to true, as we don't do any writing or editing of the file. This also allows the program to explicitly close the 
                    workbook object and avoid any I/O Operation errors being raised.
    """

    wb_obj = load_workbook(filename=workbook, keep_vba=keep_vba, data_only=data_only, read_only=True)
    wsheet = wb_obj[sheetname]

    data_key = []
    for value in wsheet.iter_rows(values_only=True):
        for key in value:
            try:
                key_strip = key.strip()
                data_key.append(key_strip)
            except AttributeError as e:
                print(f"AttributeError on key: {key}, {e}")
                data_key.append(key)
        break

    rows = []
    for rows_list in wsheet.iter_rows(values_only=True, min_row=2):
        row = []
        for values in rows_list:
            row.append(values)
        results = dict(zip(data_key, row))
        rows.append(results)

    wb_obj.close()
    
    return Result(host=task.host, result=rows)
