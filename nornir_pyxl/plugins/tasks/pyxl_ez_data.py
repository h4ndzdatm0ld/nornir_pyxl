from nornir.core.task import Result, Task
from openpyxl import load_workbook
from nornir_pyxl.plugins.tasks.helpers import open_excel_wb
import sys


def pyxl_ez_data(
    task: Task,
    workbook: str,
    sheetname: str,
) -> Result:

    """Loads a specific sheet from a workbook(xlsx file).

    Creates a list of dictionaries using the first row as the keys.

    Arguments:

        workbook: Full path to .xlsx file\n
        sheetname: Worksheet Name\n

    Examples:

        nr.run(task=pyxl_ez_data, workbookfile="example-wb.xlsx',
               sheetname='ip_data')
    Returns:

        Result object with the following attributes set:
        * result (''list''): list of dictionaries with
        data from the specific worksheet within the workbook.

    Notes:

        read_only: This is hardcoded set to true, as we don't do any writing or
        editing of the file. This also allows the program to explicitly close
        the workbook object and avoid any I/O Operation errors being raised.
    """
    wsheet = open_excel_wb(workbook, sheetname)

    data_key = []
    for value in wsheet.iter_rows(values_only=True):
        for key in value:
            try:
                key_strip = key.strip()
                data_key.append(key_strip)
            except AttributeError as e:
                print(f"AttributeError on key: {key}, {e}")
                data_key.append(key)
        break

    rows = []
    for rows_list in wsheet.iter_rows(values_only=True, min_row=2):
        row = []
        for values in rows_list:
            row.append(values)
        results = dict(zip(data_key, row))
        rows.append(results)

    return Result(host=task.host, result=rows)
